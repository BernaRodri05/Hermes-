"""
HERMES V1 - Envío automático de mensajes de WhatsApp
Autor: Berna - 2025
Con procesador de Excel/CSV integrado
"""

import subprocess
import time
import random
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import threading
from datetime import datetime, timedelta
import sys
import csv
import io
import urllib.parse


def _clamp(value):
    return max(0, min(255, int(value)))


def lighten_color(color, factor=0.1):
    """Lighten a HEX color by the given factor (0-1)."""
    color = color.lstrip('#')
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)

    r = _clamp(r + (255 - r) * factor)
    g = _clamp(g + (255 - g) * factor)
    b = _clamp(b + (255 - b) * factor)

    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"


def darken_color(color, factor=0.1):
    """Darken a HEX color by the given factor (0-1)."""
    color = color.lstrip('#')
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)

    r = _clamp(r * (1 - factor))
    g = _clamp(g * (1 - factor))
    b = _clamp(b * (1 - factor))

    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"


def create_rounded_rectangle(canvas, x1, y1, x2, y2, radius, **kwargs):
    """Draw a rounded rectangle on the given canvas."""
    radius = max(0, min(radius, (x2 - x1) / 2, (y2 - y1) / 2))
    if radius == 0:
        return canvas.create_rectangle(x1, y1, x2, y2, **kwargs)

    points = [
        x1 + radius, y1,
        x2 - radius, y1,
        x2, y1,
        x2, y1 + radius,
        x2, y2 - radius,
        x2, y2,
        x2 - radius, y2,
        x1 + radius, y2,
        x1, y2,
        x1, y2 - radius,
        x1, y1 + radius,
        x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, **kwargs)


class ShadowButton:
    """Custom button with a floating shadow frame and neumorphic styling."""

    def __init__(
        self,
        parent,
        text,
        command,
        base_bg,
        active_bg,
        text_color='#ffffff',
        font=('Inter', 13, 'bold'),
        shadow_color=None,
        shadow_offset=(4, 6),
        hover_bg=None,
        disabled_bg='#d1d5db',
        disabled_fg='#9ca3af',
        cursor='hand2',
        padding=(24, 14),
        corner_radius=24
    ):
        self.parent = parent
        self.parent_bg = parent.cget('bg') if hasattr(parent, 'cget') else '#ffffff'
        self.wrapper = tk.Frame(parent, bg=self.parent_bg)
        self.base_bg = base_bg
        self.text_color = text_color
        self.active_bg = active_bg
        self._custom_hover = hover_bg is not None
        self._custom_shadow = shadow_color is not None
        self.hover_bg = hover_bg or lighten_color(base_bg, 0.08)
        self.shadow_color = shadow_color or darken_color(self.base_bg, 0.25)
        self.disabled_bg = disabled_bg
        self.disabled_fg = disabled_fg
        self.cursor = cursor
        self.command = command
        self.corner_radius = corner_radius
        self.shadow_offset = shadow_offset
        self.padding = padding
        self.font = font
        self.font_obj = tkfont.Font(font=font)
        self.text = text

        self.button_height = self.font_obj.metrics('linespace') + (self.padding[1] * 2)
        self.min_width = self.font_obj.measure(self.text) + (self.padding[0] * 2)

        outline_color = darken_color(base_bg, 0.18)
        self.outline_color = outline_color

        self.canvas = tk.Canvas(
            self.wrapper,
            bg=self.parent_bg,
            bd=0,
            highlightthickness=0,
            cursor=self.cursor,
        )
        self.canvas.pack(fill=tk.X, expand=True)
        self.canvas.configure(height=self._get_canvas_height())

        self.button_id = None
        self.shadow_id = None
        self.text_id = None
        self._button_bbox = (0, 0, 0, 0)
        self._current_bg = self.base_bg
        self._current_width = self.min_width + self.shadow_offset[0] + 2

        self.canvas.bind('<Configure>', self._on_canvas_configure)
        self.canvas.bind('<Enter>', self._on_enter)
        self.canvas.bind('<Leave>', self._on_leave)
        self.canvas.bind('<ButtonPress-1>', self._on_press)
        self.canvas.bind('<ButtonRelease-1>', self._on_release)

        self.state = tk.NORMAL
        self._pressed = False

        self._redraw()

    # Geometry management proxies
    def pack(self, *args, **kwargs):
        return self.wrapper.pack(*args, **kwargs)

    def grid(self, *args, **kwargs):
        return self.wrapper.grid(*args, **kwargs)

    def place(self, *args, **kwargs):
        return self.wrapper.place(*args, **kwargs)

    # Configuration methods
    def configure(self, **kwargs):
        if 'text' in kwargs:
            self.text = kwargs.pop('text')
            self.min_width = self.font_obj.measure(self.text) + (self.padding[0] * 2)
            if self.text_id:
                self.canvas.itemconfigure(self.text_id, text=self.text)
            self._redraw()
        if 'command' in kwargs:
            self.command = kwargs.pop('command')
        if 'state' in kwargs:
            self._set_state(kwargs.pop('state'))
        if 'fg' in kwargs:
            self.text_color = kwargs['fg']
            if self.state == tk.NORMAL and self.text_id:
                self.canvas.itemconfigure(self.text_id, fill=self.text_color)
            kwargs.pop('fg')
        if 'bg' in kwargs:
            self.base_bg = kwargs['bg']
            if not self._custom_hover:
                self.hover_bg = lighten_color(self.base_bg, 0.08)
            if not self._custom_shadow:
                self.shadow_color = darken_color(self.base_bg, 0.25)
            self._apply_shadow_fill()
            if self.state == tk.NORMAL:
                self._set_bg(self.base_bg, update_outline=True)
            else:
                self.outline_color = darken_color(self.base_bg, 0.18)
            kwargs.pop('bg')
        if 'font' in kwargs:
            self.font = kwargs['font']
            self.font_obj = tkfont.Font(font=self.font)
            self.button_height = self.font_obj.metrics('linespace') + (self.padding[1] * 2)
            self.min_width = self.font_obj.measure(self.text) + (self.padding[0] * 2)
            self.canvas.configure(height=self._get_canvas_height())
            if self.text_id:
                self.canvas.itemconfigure(self.text_id, font=self.font)
            kwargs.pop('font')
        if kwargs:
            if self.text_id:
                self.canvas.itemconfigure(self.text_id, **kwargs)

    config = configure

    def _set_state(self, state):
        self.state = state
        if state == tk.DISABLED:
            self._set_bg(self.disabled_bg, update_outline=True)
            if self.text_id:
                self.canvas.itemconfigure(self.text_id, fill=self.disabled_fg)
            self._apply_shadow_fill()
            self.canvas.configure(cursor='arrow')
        else:
            self._set_bg(self.base_bg, update_outline=True)
            if self.text_id:
                self.canvas.itemconfigure(self.text_id, fill=self.text_color)
            self._apply_shadow_fill()
            self.canvas.configure(cursor=self.cursor)

    def _set_bg(self, color, update_outline=False):
        self._current_bg = color
        if update_outline:
            self.outline_color = darken_color(color, 0.18)
        if self.button_id:
            self.canvas.itemconfigure(self.button_id, fill=color, outline=self.outline_color)

    def _apply_shadow_fill(self):
        if not self.shadow_id:
            return
        fill = self.shadow_color if self.state == tk.NORMAL else darken_color(self.disabled_bg, 0.1)
        self.canvas.itemconfigure(self.shadow_id, fill=fill)

    # Event handlers
    def _on_enter(self, _event):
        if self.state != tk.NORMAL:
            return
        self._set_bg(self.hover_bg)

    def _on_leave(self, _event):
        if self.state != tk.NORMAL:
            return
        self._pressed = False
        self._set_bg(self.base_bg)

    def _on_press(self, _event):
        if self.state != tk.NORMAL:
            return
        self._pressed = True
        self._set_bg(self.active_bg)

    def _on_release(self, event):
        if self.state != tk.NORMAL or not self._pressed:
            return
        self._pressed = False
        x1, y1, x2, y2 = self._button_bbox
        inside = x1 <= event.x <= x2 and y1 <= event.y <= y2
        if inside and callable(self.command):
            self.command()
        self._set_bg(self.hover_bg if inside else self.base_bg)

    # Internal drawing helpers
    def _get_canvas_height(self):
        return self.button_height + self.shadow_offset[1] + 2

    def _on_canvas_configure(self, event):
        self._current_width = event.width
        self._redraw()

    def _redraw(self):
        width = max(self._current_width, self.min_width + self.shadow_offset[0] + 2)
        button_width = max(width - self.shadow_offset[0], self.min_width)
        button_height = self.button_height
        radius = min(self.corner_radius, button_height / 2)

        x1 = 0
        y1 = 0
        x2 = x1 + button_width
        y2 = y1 + button_height

        shadow_x1 = x1 + self.shadow_offset[0]
        shadow_y1 = y1 + self.shadow_offset[1]
        shadow_x2 = shadow_x1 + button_width
        shadow_y2 = shadow_y1 + button_height

        self.canvas.delete('all')

        self.shadow_id = create_rounded_rectangle(
            self.canvas,
            shadow_x1,
            shadow_y1,
            shadow_x2,
            shadow_y2,
            radius,
            fill=self.shadow_color if self.state == tk.NORMAL else darken_color(self.disabled_bg, 0.1),
            outline=''
        )

        self.button_id = create_rounded_rectangle(
            self.canvas,
            x1,
            y1,
            x2,
            y2,
            radius,
            fill=self._current_bg,
            outline=self.outline_color,
            width=1
        )

        text_color = self.text_color if self.state == tk.NORMAL else self.disabled_fg
        self.text_id = self.canvas.create_text(
            x1 + button_width / 2,
            y1 + button_height / 2,
            text=self.text,
            fill=text_color,
            font=self.font
        )
        self.canvas.tag_raise(self.text_id, self.button_id)
        self._button_bbox = (x1, y1, x2, y2)
        self._apply_shadow_fill()


# Verificar dependencias
try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("\n" + "="*50)
    print("ERROR: Falta instalar dependencias")
    print("="*50)
    print("\nPor favor ejecuta INSTALAR.bat primero")
    print("\nPresiona Enter para salir...")
    input()
    sys.exit(1)

try:
    from PIL import Image, ImageTk
except ImportError:
    print("\n" + "="*50)
    print("ERROR: Falta instalar Pillow")
    print("="*50)
    print("\nPor favor ejecuta INSTALAR.bat primero")
    print("\nPresiona Enter para salir...")
    input()
    sys.exit(1)

class Hermes:
    def __init__(self, root):
        self.root = root
        self.root.title("HERMES V1")
        self.root.minsize(1500, 900)
        self.root.configure(bg="#f8f9fa")
        
        # Variables
        self.adb_path = tk.StringVar(value="")
        self.delay_min = tk.IntVar(value=10)
        self.delay_max = tk.IntVar(value=15)
        self.wait_after_open = tk.IntVar(value=15)
        self.wait_after_first_enter = tk.IntVar(value=10)
        
        self.excel_file = ""
        self.links = []
        self.devices = []
        self.is_running = False
        self.is_paused = False
        self.should_stop = False
        self.pause_lock = threading.Lock()
        
        self.total_messages = 0
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = None

        # Datos manuales
        self.manual_numbers = []
        self.manual_messages = []
        self.manual_mode = False
        self.manual_loops = 1

        # Variables del procesador
        self.raw_data = []
        self.columns = []
        self.selected_columns = []
        self.phone_columns = []

        # Fidelizado
        self.fidelizado_unlocked = False
        self.fidelizado_wrapper = None
        self.fidelizado_trigger = None
        self.fidelizado_unlock_btn = None
        
        # Colores de Hermes
        self.colors = {
            'blue': '#4285F4',      # Azul Google
            'green': '#1DB954',     # Verde Spotify
            'orange': '#FDB913',    # Naranja/Amarillo
            'bg': '#f8f9fa',
            'text': '#202124',
            'text_light': '#5f6368',
            'action_detect': '#2563EB',
            'action_excel': '#F97316',
            'action_fidelizador': '#111827',
            'action_start': '#16A34A',
            'action_pause': '#FB923C',
            'action_cancel': '#DC2626',
        }
        
        self.setup_ui()
        self.auto_detect_adb()
        
    def setup_ui(self):
        """Configurar interfaz"""
        # Header con logo y título
        header = tk.Frame(self.root, bg=self.colors['bg'], height=150)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        header_content = tk.Frame(header, bg=self.colors['bg'])
        header_content.pack(expand=True, fill=tk.X)
        
        # Logo izquierdo
        try:
            logo_left_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo_left.png')
            logo_left_img = Image.open(logo_left_path)
            logo_left_img = logo_left_img.resize((100, 100), Image.Resampling.LANCZOS)
            logo_left_photo = ImageTk.PhotoImage(logo_left_img)
            logo_left = tk.Label(header_content, image=logo_left_photo, bg=self.colors['bg'])
            logo_left.image = logo_left_photo
            logo_left.pack(side=tk.LEFT, padx=(40, 20))
        except:
            logo_left = tk.Label(header_content, text="🦶", font=('Inter', 60),
                           bg=self.colors['bg'])
            logo_left.pack(side=tk.LEFT, padx=(40, 20))
        
        # Título centrado
        title = tk.Label(header_content, text="HERMES",
                        font=('Inter', 60, 'bold'),
                        bg=self.colors['bg'], fg=self.colors['text'])
        title.pack(side=tk.LEFT, expand=True)
        
        # Logo derecho
        try:
            logo_right_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo_right.png')
            logo_right_img = Image.open(logo_right_path)
            logo_right_img = logo_right_img.resize((100, 100), Image.Resampling.LANCZOS)
            logo_right_photo = ImageTk.PhotoImage(logo_right_img)
            logo_right = tk.Label(header_content, image=logo_right_photo, bg=self.colors['bg'])
            logo_right.image = logo_right_photo
            logo_right.pack(side=tk.RIGHT, padx=(20, 40))
        except:
            logo_right = tk.Label(header_content, text="🦶", font=('Inter', 60),
                           bg=self.colors['bg'])
            logo_right.pack(side=tk.RIGHT, padx=(20, 40))
        
        # Container principal con scroll
        main_container = tk.Frame(self.root, bg=self.colors['bg'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=60, pady=(0, 40))
        
        # Canvas para scroll
        canvas = tk.Canvas(main_container, bg=self.colors['bg'], highlightthickness=0)
        self.main_canvas = canvas
        scrollbar = tk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        main = tk.Frame(canvas, bg=self.colors['bg'])
        main.grid_columnconfigure(0, weight=618, uniform='main_panels')
        main.grid_columnconfigure(1, weight=382, uniform='main_panels')
        main.grid_rowconfigure(0, weight=1)
        main.grid_rowconfigure(1, weight=1)

        self.main_layout = main
        self.main_window = canvas.create_window((0, 0), window=main, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Habilitar scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _on_canvas_configure(event):
            canvas.itemconfigure(self.main_window, width=event.width)
            self._update_main_layout(event.width)

        canvas.bind("<Configure>", _on_canvas_configure)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Dos columnas adaptables
        left = tk.Frame(main, bg=self.colors['bg'])
        right = tk.Frame(main, bg=self.colors['bg'])

        self.left_panel = left
        self.right_panel = right
        self._current_main_layout = None

        def _on_main_configure(_event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            self._update_main_layout(self._get_available_main_width())

        main.bind("<Configure>", _on_main_configure)

        self.setup_left(left)
        self.setup_right(right)

        main.update_idletasks()
        self._update_main_layout(self._get_available_main_width())
        
    def setup_left(self, parent):
        """Panel izquierdo"""
        # Configuración de Tiempo
        config_title = tk.Frame(parent, bg=self.colors['bg'])
        config_title.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(config_title, text="⚙️", font=('Inter', 20),
                bg=self.colors['bg']).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Label(config_title, text="Configuración de Tiempo",
                font=('Inter', 16, 'bold'),
                bg=self.colors['bg'], fg='#000000').pack(side=tk.LEFT)
        
        tk.Frame(parent, bg='#e0e0e0', height=1).pack(fill=tk.X, pady=(0, 25))
        
        # Settings
        settings = tk.Frame(parent, bg=self.colors['bg'])
        settings.pack(fill=tk.X, pady=(0, 40))
        
        self.create_setting(settings, "Delay entre mensajes (seg):",
                          self.delay_min, self.delay_max, 0)
        self.create_setting(settings, "Espera después de abrir (seg):",
                          self.wait_after_open, None, 1)
        self.create_setting(settings, "Espera después del 1er ENTER (seg):",
                          self.wait_after_first_enter, None, 2)
        
        # Acciones
        actions_title = tk.Frame(parent, bg=self.colors['bg'])
        actions_title.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(actions_title, text="👤", font=('Inter', 20),
                bg=self.colors['bg']).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Label(actions_title, text="Acciones",
                font=('Inter', 16, 'bold'),
                bg=self.colors['bg'], fg='#000000').pack(side=tk.LEFT)

        unlock_wrapper = tk.Frame(actions_title, bg=self.colors['bg'])
        unlock_wrapper.pack(side=tk.LEFT, padx=(12, 0))

        unlock_shadow = tk.Frame(unlock_wrapper, bg='#c8ccd5', bd=0)
        unlock_shadow.place(x=1, y=2)

        unlock_button_container = tk.Frame(unlock_wrapper, bg='#f7f9fc', bd=0)
        unlock_button_container.pack()

        def _sync_unlock_shadow(event):
            unlock_shadow.place_configure(width=event.width, height=event.height)

        unlock_button_container.bind("<Configure>", _sync_unlock_shadow)
        unlock_shadow.lower()

        self.fidelizado_unlock_btn = tk.Button(
            unlock_button_container,
            text="🔒",
            command=self.request_fidelizado_access,
            bg='#ffffff', fg=self.colors['text'],
            font=('Inter', 11, 'bold'),
            relief=tk.RAISED, cursor='hand2',
            activebackground='#e5e7eb',
            bd=1,
            padx=6,
            pady=2,
            highlightthickness=1,
            highlightbackground='#c8ccd5',
            highlightcolor='#c8ccd5'
        )
        self.fidelizado_unlock_btn.pack()
        
        tk.Frame(parent, bg='#e0e0e0', height=1).pack(fill=tk.X, pady=(0, 25))
        
        # Botones
        actions = tk.Frame(parent, bg=self.colors['bg'])
        actions.pack(fill=tk.X)
        
        # Botón 1
        btn1_container = tk.Frame(actions, bg=self.colors['bg'])
        btn1_container.pack(fill=tk.X, pady=(0, 15))
        
        num1 = tk.Label(btn1_container, text="1",
                       font=('Inter', 20, 'bold'),
                       bg='#e8eaed', fg=self.colors['text'],
                       width=3, height=1)
        num1.pack(side=tk.LEFT, padx=(0, 15))
        
        self.btn_detect = ShadowButton(
            btn1_container,
            text="🔍  Detectar Dispositivos",
            command=self.detect_devices,
            base_bg=self.colors['action_detect'],
            active_bg=darken_color(self.colors['action_detect'], 0.18),
            text_color='white'
        )
        self.btn_detect.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Botón 2
        btn2_container = tk.Frame(actions, bg=self.colors['bg'])
        btn2_container.pack(fill=tk.X, pady=(0, 15))
        
        num2 = tk.Label(btn2_container, text="2",
                       font=('Inter', 20, 'bold'),
                       bg='#e8eaed', fg=self.colors['text'],
                       width=3, height=1)
        num2.pack(side=tk.LEFT, padx=(0, 15))
        
        self.btn_load = ShadowButton(
            btn2_container,
            text="📄  Cargar y Procesar Excel",
            command=self.load_and_process_excel,
            base_bg=self.colors['action_excel'],
            active_bg=darken_color(self.colors['action_excel'], 0.18),
            text_color='white'
        )
        self.btn_load.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Sección Fidelizado
        fidelizado_section = tk.Frame(actions, bg=self.colors['bg'])
        fidelizado_section.pack(fill=tk.X, pady=(0, 20))

        secret_trigger_wrapper = tk.Frame(fidelizado_section, bg=self.colors['bg'])
        secret_trigger_wrapper.pack(fill=tk.X, padx=(48, 0))

        self.fidelizado_trigger = ShadowButton(
            secret_trigger_wrapper,
            text="📱  Fidelizado",
            command=self.handle_fidelizado_access,
            base_bg=self.colors['action_fidelizador'],
            active_bg=darken_color(self.colors['action_fidelizador'], 0.18),
            text_color='#ffffff',
            font=('Inter', 13, 'bold'),
            padding=(24, 14),
            corner_radius=18
        )
        self.fidelizado_trigger.pack(fill=tk.X, padx=12, pady=12)

        self.fidelizado_trigger.configure(state=tk.DISABLED)
        secret_trigger_wrapper.pack_forget()

        self.fidelizado_wrapper = secret_trigger_wrapper

        # Botón 3
        btn3_container = tk.Frame(actions, bg=self.colors['bg'])
        btn3_container.pack(fill=tk.X, pady=(0, 25))

        num3 = tk.Label(btn3_container, text="3",
                       font=('Inter', 20, 'bold'),
                       bg='#e8eaed', fg=self.colors['text'],
                       width=3, height=1)
        num3.pack(side=tk.LEFT, padx=(0, 15))
        
        self.btn_start = ShadowButton(
            btn3_container,
            text="▶  INICIAR ENVÍO",
            command=self.start_sending,
            base_bg=self.colors['action_start'],
            active_bg=darken_color(self.colors['action_start'], 0.18),
            text_color='white'
        )
        self.btn_start.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Controles
        controls = tk.Frame(actions, bg=self.colors['bg'])
        controls.pack(fill=tk.X, pady=(10, 0))
        
        self.btn_pause = ShadowButton(
            controls,
            text="⏸  PAUSAR",
            command=self.pause_sending,
            base_bg=self.colors['action_pause'],
            active_bg=darken_color(self.colors['action_pause'], 0.18),
            text_color='#ffffff',
            font=('Inter', 12, 'bold'),
            disabled_bg='#e5e7eb',
            disabled_fg='#9ca3af',
            padding=(22, 12)
        )
        self.btn_pause.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.btn_pause.configure(state=tk.DISABLED)

        self.btn_stop = ShadowButton(
            controls,
            text="⏹  CANCELAR",
            command=self.stop_sending,
            base_bg=self.colors['action_cancel'],
            active_bg=darken_color(self.colors['action_cancel'], 0.18),
            text_color='#ffffff',
            font=('Inter', 12, 'bold'),
            disabled_bg='#e5e7eb',
            disabled_fg='#9ca3af',
            padding=(22, 12)
        )
        self.btn_stop.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0))
        self.btn_stop.configure(state=tk.DISABLED)
        
    def setup_right(self, parent):
        """Panel derecho"""
        title = tk.Frame(parent, bg=self.colors['bg'])
        title.pack(fill=tk.X, pady=(0, 25))

        tk.Label(title, text="✓", font=('Inter', 20),
                bg=self.colors['bg'], fg=self.colors['green']).pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Label(title, text="Estado y Progreso",
                font=('Inter', 16, 'bold'),
                bg=self.colors['bg'], fg='#000000').pack(side=tk.LEFT)
        
        tk.Frame(parent, bg='#e0e0e0', height=1).pack(fill=tk.X, pady=(0, 30))
        
        # Stats
        stats = tk.Frame(parent, bg=self.colors['bg'])
        stats.pack(fill=tk.BOTH, expand=True, pady=(0, 35))
        self.stats_frame = stats
        
        self.create_stat(stats, "Total", "0", self.colors['blue'], 0)
        self.create_stat(stats, "Enviados", "0", self.colors['green'], 1)
        self.create_stat(stats, "Progreso", "0%", self.colors['orange'], 2)
        
        tk.Label(parent, text="Progreso general",
                font=('Inter', 12, 'bold'),
                bg=self.colors['bg'], fg=self.colors['text']).pack(anchor='w', pady=(0, 5))
        
        self.progress_label = tk.Label(parent, text="--/--",
                                      font=('Inter', 20, 'bold'),
                                      bg=self.colors['bg'], fg=self.colors['text'])
        self.progress_label.pack(anchor='w', pady=(0, 10))
        
        bar_bg = tk.Frame(parent, bg='#e0e0e0', height=6)
        bar_bg.pack(fill=tk.X, pady=(0, 20))
        
        self.progress_bar = tk.Frame(bar_bg, bg=self.colors['green'], height=6)
        self.progress_bar.place(x=0, y=0, relwidth=0, relheight=1)
        
        time_title = tk.Frame(parent, bg=self.colors['bg'])
        time_title.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(time_title, text="⏱", font=('Inter', 14),
                bg=self.colors['bg']).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Label(time_title, text="Tiempo",
                font=('Inter', 12, 'bold'),
                bg=self.colors['bg'], fg=self.colors['text']).pack(side=tk.LEFT)
        
        self.time_elapsed = tk.Label(parent, text="Transcurrido: --:--:--",
                                     font=('Inter', 10),
                                     bg=self.colors['bg'], fg=self.colors['text_light'])
        self.time_elapsed.pack(anchor='w', pady=2)
        
        self.time_remaining = tk.Label(parent, text="Restante: --:--:--",
                                       font=('Inter', 10),
                                       bg=self.colors['bg'], fg=self.colors['text_light'])
        self.time_remaining.pack(anchor='w', pady=(2, 20))
        
        log_title = tk.Frame(parent, bg=self.colors['bg'])
        log_title.pack(fill=tk.X, pady=(0, 12))
        
        tk.Label(log_title, text="▶", font=('Inter', 14),
                bg=self.colors['bg']).pack(side=tk.LEFT, padx=(0, 8))
        
        tk.Label(log_title, text="Registro de actividad",
                font=('Inter', 16, 'bold'),
                bg=self.colors['bg'], fg='#000000').pack(side=tk.LEFT)
        
        log_container = tk.Frame(parent, bg='#1e1e1e')
        log_container.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        self.log_container = log_container
        
        scrollbar = tk.Scrollbar(log_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(log_container, bg='#1e1e1e', fg='#00d9ff',
                               font=('Consolas', 10), relief=tk.FLAT,
                               yscrollcommand=scrollbar.set, bd=0,
                               padx=15, pady=15)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)
        
        self.log_text.tag_config('success', foreground='#00d9ff')
        self.log_text.tag_config('error', foreground='#ff5555')
        self.log_text.tag_config('warning', foreground='#ffaa00')
        self.log_text.tag_config('info', foreground='#00d9ff')
        
        self.log("✓ HERMES V1 iniciado", 'success')
        self.log("ℹ Sigue los pasos 1, 2, 3 y 4", 'info')
        self.log("✓ ADB detectado", 'success')

    def _get_available_main_width(self):
        """Obtener el ancho disponible para el layout principal."""
        width = 0
        if hasattr(self, 'main_canvas'):
            width = self.main_canvas.winfo_width()
        if not width and hasattr(self, 'root'):
            width = self.root.winfo_width()
        return width

    def _update_main_layout(self, width=None):
        """Actualizar disposición principal entre columnas y modo apilado"""
        if not hasattr(self, 'left_panel') or not hasattr(self, 'right_panel'):
            return

        if not width:
            width = self._get_available_main_width()

        threshold = 1100
        mode = 'stacked' if width and width < threshold else 'columns'

        if self._current_main_layout == mode:
            return

        if mode == 'columns':
            if hasattr(self, 'main_layout'):
                self.main_layout.grid_columnconfigure(0, weight=618, uniform='main_panels', minsize=0)
                self.main_layout.grid_columnconfigure(1, weight=382, uniform='main_panels', minsize=0)
            self.left_panel.grid(row=0, column=0, sticky='nsew', padx=(0, 20), pady=(0, 0))
            self.right_panel.grid(row=0, column=1, sticky='nsew', padx=(20, 0), pady=(0, 0))
            if hasattr(self, 'stats_frame'):
                self.stats_frame.pack_configure(pady=(0, 35))
            if hasattr(self, 'log_container'):
                self.log_container.pack_configure(pady=(0, 10))
        else:
            if hasattr(self, 'main_layout'):
                self.main_layout.grid_columnconfigure(0, weight=1, uniform='main_panels', minsize=0)
                self.main_layout.grid_columnconfigure(1, weight=1, uniform='main_panels', minsize=0)
            self.left_panel.grid(row=0, column=0, columnspan=2, sticky='nsew', padx=0, pady=(0, 30))
            self.right_panel.grid(row=1, column=0, columnspan=2, sticky='nsew', padx=0, pady=(0, 0))
            if hasattr(self, 'stats_frame'):
                self.stats_frame.pack_configure(pady=(0, 20))
            if hasattr(self, 'log_container'):
                self.log_container.pack_configure(pady=(0, 20))

        self._current_main_layout = mode
        
    def create_stat(self, parent, label, value, color, col):
        """Crear caja de estadística"""
        box = tk.Frame(parent, bg=color, bd=0, relief=tk.FLAT)
        box.grid(row=0, column=col, sticky='nsew', padx=8)
        parent.grid_columnconfigure(col, weight=1)
        
        tk.Label(box, text=label, bg=color, fg='white',
                font=('Inter', 13)).pack(pady=(22, 8))
        
        val_label = tk.Label(box, text=value, bg=color, fg='white',
                            font=('Inter', 48, 'bold'))
        val_label.pack(pady=(0, 22))
        
        if label == "Total":
            self.stat_total = val_label
        elif label == "Enviados":
            self.stat_sent = val_label
        elif label == "Progreso":
            self.stat_progress = val_label
            
    def create_setting(self, parent, label, var1, var2, row):
        """Crear fila de configuración"""
        tk.Label(parent, text=label,
                font=('Inter', 12),
                bg=self.colors['bg'], fg=self.colors['text']).grid(
                    row=row, column=0, sticky='w', pady=15)
        
        controls = tk.Frame(parent, bg=self.colors['bg'])
        controls.grid(row=row, column=1, sticky='e', pady=15, padx=(20, 0))
        
        combo1 = ttk.Combobox(controls, textvariable=var1, width=8,
                             font=('Inter', 11), state='normal',
                             values=list(range(1, 121)))
        combo1.pack(side=tk.LEFT, padx=(0, 10))
        
        if var2:
            tk.Label(controls, text="-",
                    font=('Inter', 12),
                    bg=self.colors['bg']).pack(side=tk.LEFT, padx=(0, 10))
            
            combo2 = ttk.Combobox(controls, textvariable=var2, width=8,
                                 font=('Inter', 11), state='normal',
                                 values=list(range(1, 121)))
            combo2.pack(side=tk.LEFT)
            
    def log(self, msg, tag='info'):
        """Agregar al log"""
        ts = datetime.now().strftime("[%H:%M:%S]")
        self.log_text.insert(tk.END, f"{ts} {msg}\n", tag)
        self.log_text.see(tk.END)
        self.root.update()
        
    def update_stats(self):
        """Actualizar estadísticas"""
        self.stat_total.config(text=str(self.total_messages))
        self.stat_sent.config(text=str(self.sent_count))
        
        if self.total_messages > 0:
            prog = int((self.current_index / self.total_messages) * 100)
            self.stat_progress.config(text=f"{prog}%")
            self.progress_bar.place(relwidth=prog/100)
            self.progress_label.config(text=f"{self.current_index}/{self.total_messages}")
            
            if self.start_time and self.current_index > 0:
                elapsed = datetime.now() - self.start_time
                self.time_elapsed.config(text=f"Transcurrido: {str(elapsed).split('.')[0]}")
                
                avg = elapsed.total_seconds() / self.current_index
                rem_sec = avg * (self.total_messages - self.current_index)
                rem = timedelta(seconds=int(rem_sec))
                self.time_remaining.config(text=f"Restante: {str(rem).split('.')[0]}")
                
    def auto_detect_adb(self):
        """Detectar ADB"""
        current_dir = os.path.dirname(os.path.abspath(__file__))
        paths = [
            os.path.join(current_dir, "scrcpy-win64-v3.2", "adb.exe"),
            os.path.join(current_dir, "adb.exe"),
        ]
        
        for path in paths:
            if os.path.exists(path):
                self.adb_path.set(path)
                return
                
    def detect_devices(self):
        """Detectar dispositivos"""
        adb = self.adb_path.get()
        if not adb or not os.path.exists(adb):
            messagebox.showerror("Error", "ADB no encontrado")
            return
            
        self.log("🔍 Detectando dispositivos...", 'info')
        
        try:
            result = subprocess.run([adb, 'devices'], capture_output=True,
                                   text=True, timeout=10)
            self.devices = []
            for line in result.stdout.strip().split('\n')[1:]:
                if '\tdevice' in line:
                    self.devices.append(line.split('\t')[0])
                    
            if self.devices:
                self.log(f"✓ {len(self.devices)} dispositivo(s) encontrado(s)", 'success')
            else:
                self.log("✗ No se encontraron dispositivos", 'error')
        except Exception as e:
            self.log(f"✗ Error: {e}", 'error')
    
    def read_csv_file(self, filepath):
        """Leer archivo CSV con detección de codificación y soporte completo para emojis"""
        try:
            # Priorizar UTF-8 para emojis
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-16']
            
            for encoding in encodings:
                try:
                    with open(filepath, 'r', encoding=encoding, errors='ignore') as file:
                        sample = file.read(2048)
                        file.seek(0)
                        
                        delimiters = [';', ',', '\t', '|']
                        delimiter = ','
                        for delim in delimiters:
                            if delim in sample:
                                delimiter = delim
                                break
                        
                        reader = csv.DictReader(file, delimiter=delimiter)
                        data = []
                        for row in reader:
                            clean_row = {}
                            for key, value in row.items():
                                if key is not None:
                                    clean_key = key.strip()
                                    # Preservar emojis y caracteres especiales
                                    clean_value = value if value is not None else ''
                                    clean_row[clean_key] = clean_value
                            data.append(clean_row)
                        
                        fieldnames = [name.strip() for name in reader.fieldnames if name is not None] if reader.fieldnames else []
                        return data, fieldnames
                except:
                    continue
            
            raise Exception("No se pudo leer el archivo CSV con ninguna codificación")
        except Exception as e:
            raise Exception(f"Error al leer archivo CSV: {str(e)}")
    
    def read_excel_file(self, filepath):
        """Leer archivo Excel usando openpyxl"""
        try:
            workbook = load_workbook(filepath, data_only=True)
            sheet = workbook.active
            
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value).strip() if cell.value is not None else '')
            
            data = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        if value is None or value == '':
                            row_dict[headers[col_idx]] = ''
                        elif isinstance(value, (int, float)):
                            row_dict[headers[col_idx]] = str(value)
                        else:
                            row_dict[headers[col_idx]] = str(value)
                data.append(row_dict)
            
            return data, headers
        except Exception as e:
            raise Exception(f"Error al leer archivo Excel: {str(e)}")
    
    def load_and_process_excel(self):
        """Cargar y procesar Excel/CSV"""
        self.log("📂 Seleccionando archivo...", 'info')

        self.manual_mode = False

        file_path = filedialog.askopenfilename(
            title="Seleccionar Excel o CSV",
            filetypes=[("Excel", "*.xlsx *.xls"), ("CSV", "*.csv"), ("Todos", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            self.log("📖 Leyendo archivo...", 'info')
            
            if file_path.lower().endswith('.csv'):
                self.raw_data, self.columns = self.read_csv_file(file_path)
            else:
                self.raw_data, self.columns = self.read_excel_file(file_path)
            
            # DETECCIÓN INTELIGENTE: Verificar si ya tiene URLs
            if 'URL' in self.columns or 'url' in self.columns:
                # Excel ya procesado con URLs
                url_col = 'URL' if 'URL' in self.columns else 'url'
                self.links = [row[url_col] for row in self.raw_data if row.get(url_col)]
                
                if self.links:
                    self.total_messages = len(self.links)
                    self.update_stats()
                    self.log(f"✓ Excel con URLs cargado: {len(self.links)} URLs detectados", 'success')
                    messagebox.showinfo("Excel Cargado", 
                                      f"Se detectaron {len(self.links)} URLs de WhatsApp\n\n"
                                      "Puedes iniciar el envío directamente.")
                    return
            
            # Excel original sin URLs - procesar normalmente
            self.phone_columns = [col for col in self.columns if col and 'telefono' in col.lower()]
            
            if not self.phone_columns:
                messagebox.showerror("Error", "No se encontraron columnas de teléfono en el archivo")
                return
            
            self.log(f"✓ Archivo leído: {len(self.raw_data)} filas, {len(self.columns)} columnas", 'success')
            self.log(f"✓ Columnas de teléfono: {', '.join(self.phone_columns)}", 'success')
            
            self.open_processor_window(file_path)
            
        except Exception as e:
            self.log(f"✗ Error al leer archivo: {e}", 'error')
            messagebox.showerror("Error", f"Error al leer archivo: {e}")

    def _show_fidelizado_trigger(self):
        """Mostrar el botón Fidelizado respetando la sombra"""
        if self.fidelizado_wrapper and not self.fidelizado_wrapper.winfo_manager():
            self.fidelizado_wrapper.pack(fill=tk.X, padx=(48, 0))

    def _prompt_fidelizado_password(self):
        """Ventana personalizada para solicitar la contraseña de Fidelizado"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Acceso Fidelizado")
        dialog.configure(bg=self.colors['bg'])
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, False)

        dialog.update_idletasks()
        width, height = 360, 240
        root_width = max(self.root.winfo_width(), width)
        root_height = max(self.root.winfo_height(), height)
        x = self.root.winfo_rootx() + (root_width // 2) - (width // 2)
        y = self.root.winfo_rooty() + (root_height // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")

        card = tk.Frame(dialog, bg='#ffffff', bd=0, relief=tk.FLAT)
        card.pack(fill=tk.BOTH, expand=True, padx=24, pady=24)

        header = tk.Frame(card, bg='#ffffff')
        header.pack(fill=tk.X, pady=(0, 18))

        icon_wrapper = tk.Frame(header, bg='#e4e7ef', width=54, height=54)
        icon_wrapper.pack(side=tk.LEFT)
        icon_wrapper.pack_propagate(False)
        tk.Label(icon_wrapper, text="🔐", font=('Inter', 24), bg='#e4e7ef').pack(expand=True, fill=tk.BOTH)

        tk.Label(
            header,
            text="Desbloquear Fidelizado",
            font=('Inter', 16, 'bold'),
            bg='#ffffff', fg=self.colors['text']
        ).pack(anchor='w', padx=(16, 0))

        tk.Label(
            card,
            text="Ingresa la contraseña para acceder a las funciones Fidelizado",
            font=('Inter', 11),
            bg='#ffffff', fg=self.colors['text_light'],
            wraplength=280,
            justify='left'
        ).pack(anchor='w')

        entry_container = tk.Frame(card, bg='#ffffff')
        entry_container.pack(fill=tk.X, pady=(18, 10))

        entry_wrapper = tk.Frame(entry_container, bg='#ffffff', bd=0, relief=tk.FLAT)
        entry_wrapper.pack(fill=tk.X)

        entry_shadow = tk.Frame(entry_wrapper, bg='#c8ccd5', bd=0)
        entry_shadow.place(relx=0, rely=0, relwidth=1, relheight=1, x=2, y=2)

        entry_body = tk.Frame(entry_wrapper, bg='#f8f9fc', bd=0, relief=tk.FLAT)
        entry_body.pack(fill=tk.X)

        password_var = tk.StringVar()
        password_entry = tk.Entry(
            entry_body,
            textvariable=password_var,
            font=('Inter', 12),
            show='*',
            relief=tk.FLAT,
            bg='#f8f9fc',
            fg=self.colors['text'],
            insertbackground=self.colors['text']
        )
        password_entry.pack(fill=tk.X, padx=14, pady=12)
        password_entry.focus_set()

        result = {'value': None}

        buttons = tk.Frame(card, bg='#ffffff')
        buttons.pack(fill=tk.X, pady=(12, 0))

        def close_dialog(value=None):
            result['value'] = value
            dialog.destroy()

        def submit(event=None):
            close_dialog(password_var.get().strip())

        def cancel(event=None):
            close_dialog(None)

        primary_btn = tk.Button(
            buttons,
            text="Desbloquear",
            command=submit,
            bg=self.colors['blue'],
            fg='white',
            font=('Inter', 11, 'bold'),
            relief=tk.FLAT,
            activebackground='#3367D6',
            cursor='hand2',
            padx=14,
            pady=8
        )
        primary_btn.pack(side=tk.RIGHT)

        secondary_btn = tk.Button(
            buttons,
            text="Cancelar",
            command=cancel,
            bg='#e5e7eb',
            fg=self.colors['text'],
            font=('Inter', 11),
            relief=tk.FLAT,
            activebackground='#d1d5db',
            cursor='hand2',
            padx=14,
            pady=8
        )
        secondary_btn.pack(side=tk.RIGHT, padx=(0, 12))

        dialog.bind('<Return>', submit)
        dialog.bind('<Escape>', cancel)
        dialog.protocol('WM_DELETE_WINDOW', cancel)

        dialog.wait_window()
        return result['value']

    def request_fidelizado_access(self):
        """Solicitar acceso protegido para Fidelizado"""
        if self.fidelizado_unlocked:
            self._show_fidelizado_trigger()
            if self.fidelizado_trigger:
                self.fidelizado_trigger.configure(state=tk.NORMAL)
            if hasattr(self, 'fidelizado_unlock_btn') and self.fidelizado_unlock_btn:
                self.fidelizado_unlock_btn.configure(text="🔓")
            return

        password = self._prompt_fidelizado_password()

        if password is None:
            return

        if password != "feli2109":
            messagebox.showerror("-beta-", "Contraseña incorrecta")
            return

        self.fidelizado_unlocked = True
        if self.fidelizado_trigger:
            self.fidelizado_trigger.configure(state=tk.NORMAL)
        if hasattr(self, 'fidelizado_unlock_btn') and self.fidelizado_unlock_btn:
            self.fidelizado_unlock_btn.configure(text="🔓")
        self._show_fidelizado_trigger()

    def handle_fidelizado_access(self):
        """Acceso unificado a la ventana de Fidelizado"""
        if not self.fidelizado_unlocked:
            return
        self.open_manual_input_window()

    def open_manual_input_window(self):
        """Ventana para ingresar números y mensajes Fidelizado"""
        manual_window = tk.Toplevel(self.root)
        manual_window.title("HERMES V1 - Fidelizado")
        manual_window.geometry("700x650")
        manual_window.configure(bg=self.colors['bg'])
        manual_window.transient(self.root)
        manual_window.grab_set()

        content = tk.Frame(manual_window, bg='white', bd=0, relief=tk.FLAT)
        content.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)

        header = tk.Label(content,
                          text="📱 Fidelizado: carga de números y mensajes",
                          font=('Inter', 18, 'bold'),
                          bg='white', fg=self.colors['text'])
        header.pack(anchor='w', pady=(0, 20))

        numbers_frame = tk.Frame(content, bg='white')
        numbers_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        tk.Label(numbers_frame,
                 text="Pega los números (uno por línea, sin prefijo +549):",
                 font=('Inter', 11, 'bold'),
                 bg='white', fg=self.colors['text']).pack(anchor='w', pady=(0, 8))

        numbers_text = scrolledtext.ScrolledText(numbers_frame,
                                                 height=12,
                                                 font=('Inter', 11),
                                                 relief=tk.SOLID,
                                                 bd=1,
                                                 wrap=tk.WORD)
        numbers_text.pack(fill=tk.BOTH, expand=True)
        numbers_text.focus_set()

        if self.manual_numbers:
            numbers_text.insert('1.0', "\n".join(self.manual_numbers))

        controls_frame = tk.Frame(content, bg='white')
        controls_frame.pack(fill=tk.X, pady=(0, 20))

        loops_var = tk.IntVar(value=max(1, self.manual_loops))

        loops_container = tk.Frame(controls_frame, bg='white')
        loops_container.pack(side=tk.LEFT, anchor='w')

        tk.Label(loops_container,
                 text="Repetir bucle de números:",
                 font=('Inter', 11, 'bold'),
                 bg='white', fg=self.colors['text']).pack(anchor='w')

        try:
            loops_input = ttk.Spinbox(loops_container, from_=1, to=9999,
                                      textvariable=loops_var, width=10, justify='center')
        except AttributeError:
            loops_input = tk.Spinbox(loops_container, from_=1, to=9999,
                                     textvariable=loops_var, width=10, justify='center')
        loops_input.pack(anchor='w', pady=(6, 0))

        message_controls = tk.Frame(controls_frame, bg='white')
        message_controls.pack(side=tk.RIGHT, anchor='e')

        messages_count_var = tk.StringVar()
        if self.manual_messages:
            messages_count_var.set(f"{len(self.manual_messages)} mensajes cargados")
        else:
            messages_count_var.set("0 mensajes cargados")

        def load_messages_from_file():
            file_path = filedialog.askopenfilename(
                title="Seleccionar archivo de mensajes (.txt)",
                filetypes=[("Texto", "*.txt"), ("Todos", "*.*")]
            )

            if not file_path:
                return

            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f.read().splitlines() if line.strip()]

                if not lines:
                    messagebox.showerror("Error", "El archivo no contiene mensajes válidos.")
                    return

                self.manual_messages = lines
                messages_count_var.set(f"{len(self.manual_messages)} mensajes cargados")
                self.log(f"✓ {len(self.manual_messages)} mensajes Fidelizado cargados", 'success')

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo leer el archivo: {e}")

        ttk.Button(message_controls,
                   text="📂 Cargar mensajes (.txt)",
                   command=load_messages_from_file).pack(anchor='e')

        tk.Label(message_controls,
                 textvariable=messages_count_var,
                 font=('Inter', 10),
                 bg='white', fg=self.colors['text_light']).pack(anchor='e', pady=(6, 0))

        buttons_frame = tk.Frame(content, bg='white')
        buttons_frame.pack(fill=tk.X, pady=(10, 0))

        def close_window():
            manual_window.grab_release()
            manual_window.destroy()

        manual_window.protocol("WM_DELETE_WINDOW", close_window)

        def confirm_manual_data():
            raw_numbers = numbers_text.get('1.0', tk.END).splitlines()
            cleaned_numbers = []

            for raw in raw_numbers:
                stripped = raw.strip()
                if not stripped:
                    continue

                normalized = ''.join(stripped.split())

                if normalized.startswith('+549'):
                    messagebox.showerror("Error", "Los números no deben incluir el prefijo +549.")
                    return

                if normalized.startswith('+'):
                    normalized = normalized[1:]

                if not normalized.isdigit():
                    messagebox.showerror("Error", f"Número inválido: {stripped}")
                    return

                cleaned_numbers.append(normalized)

            if not cleaned_numbers:
                messagebox.showerror("Error", "Ingresa al menos un número válido.")
                return

            if not self.manual_messages:
                messagebox.showerror("Error", "Carga un archivo .txt con los mensajes a enviar.")
                return

            try:
                loops_value = int(loops_var.get())
            except (TypeError, ValueError, tk.TclError):
                loops_value = 1

            if loops_value < 1:
                loops_value = 1

            links = self.generate_manual_links(cleaned_numbers, self.manual_messages, loops_value)

            if not links:
                messagebox.showerror("Error", "No se pudieron generar enlaces con los datos proporcionados.")
                return

            self.manual_numbers = cleaned_numbers
            self.manual_loops = loops_value
            self.links = links
            self.manual_mode = True
            self.total_messages = len(self.links)
            self.update_stats()

            self.log(
                f"✓ Lista Fidelizado cargada: {len(self.manual_numbers)} números, "
                f"{len(self.manual_messages)} mensajes, {len(self.links)} URLs generadas",
                'success'
            )

            messagebox.showinfo("Lista Fidelizado",
                                f"Se generaron {len(self.links)} mensajes a partir de la lista Fidelizado.")

            close_window()

        ttk.Button(buttons_frame, text="Cancelar", command=close_window).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(buttons_frame, text="Generar enlaces", command=confirm_manual_data).pack(side=tk.RIGHT)

    def generate_manual_links(self, numbers, messages, loops):
        """Generar URLs de WhatsApp a partir de números y mensajes manuales"""
        if not numbers or not messages:
            return []

        count = len(numbers)
        base_sequence = []
        for number in numbers:
            base_sequence.extend([number] * count)

        if not base_sequence:
            return []

        loops = max(1, loops)
        total_messages = len(messages)
        block_size = len(base_sequence)
        required_repeats = -(-total_messages // block_size)
        repeats = max(loops, required_repeats)

        full_sequence = (base_sequence * repeats)[:total_messages]

        links = []
        for number, message in zip(full_sequence, messages):
            encoded_message = urllib.parse.quote(message, safe='')
            links.append(f"https://wa.me/549{number}?text={encoded_message}")

        return links

    def open_processor_window(self, original_file):
        """Ventana de configuración con colores y tipografía de Hermes"""
        proc_window = tk.Toplevel(self.root)
        proc_window.title("HERMES V1 - Configurar Procesamiento")
        proc_window.geometry("900x750")
        proc_window.configure(bg=self.colors['bg'])
        
        # Container principal
        main_container = tk.Frame(proc_window, bg='white', relief=tk.FLAT)
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header con colores de Hermes
        header = tk.Frame(main_container, bg=self.colors['blue'], height=80)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        
        tk.Label(header, text="⚙️ Configurar Procesamiento de Datos",
                font=('Inter', 22, 'bold'),
                bg=self.colors['blue'], fg='white').pack(expand=True)
        
        # Contenido con scroll
        canvas = tk.Canvas(main_container, bg='white', highlightthickness=0)
        scrollbar = tk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Habilitar scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Habilitar scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            return "break"
        
        def _on_enter(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _on_leave(event):
            canvas.unbind_all("<MouseWheel>")
        
        canvas.bind("<Enter>", _on_enter)
        canvas.bind("<Leave>", _on_leave)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # PASO 1: Información (siempre visible)
        step1 = tk.Frame(scrollable_frame, bg='white')
        step1.pack(fill=tk.X, padx=30, pady=(20, 10))
        
        step1_header = tk.Frame(step1, bg='white')
        step1_header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(step1_header, text="Información del Archivo",
                font=('Inter', 13, 'bold'),
                bg='white', fg=self.colors['text']).pack(side=tk.LEFT)
        
        info_box = tk.Frame(step1, bg='#e8f5e9', relief=tk.SOLID, bd=1)
        info_box.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(info_box, text=f"📊 {os.path.basename(original_file)}",
                font=('Inter', 12, 'bold'),
                bg='#e8f5e9', fg='#2e7d32').pack(anchor='w', padx=20, pady=(15, 5))
        
        tk.Label(info_box, text=f"📝 Total de filas: {len(self.raw_data)}",
                font=('Inter', 11),
                bg='#e8f5e9', fg='#2e7d32').pack(anchor='w', padx=20, pady=5)
        
        tk.Label(info_box, text=f"📞 Columnas de teléfono: {', '.join(self.phone_columns)}",
                font=('Inter', 11),
                bg='#e8f5e9', fg='#2e7d32').pack(anchor='w', padx=20, pady=(5, 15))
        
        tk.Frame(step1, bg='#e0e0e0', height=2).pack(fill=tk.X, pady=15)
        
        # PASO 2: Teléfonos (desplegable mejorado)
        step2 = tk.Frame(scrollable_frame, bg='white')
        step2.pack(fill=tk.X, padx=30, pady=10)
        
        step2_toggle = tk.Frame(step2, bg='#f5f5f5', relief=tk.FLAT, bd=0, cursor='hand2')
        step2_toggle.pack(fill=tk.X)
        
        step2_header_inner = tk.Frame(step2_toggle, bg='#f5f5f5')
        step2_header_inner.pack(fill=tk.X, padx=15, pady=12)
        
        tk.Label(step2_header_inner, text="1", font=('Inter', 18, 'bold'),
                bg='#e8eaed', fg='#202124', width=2, height=1).pack(side=tk.LEFT, padx=(0, 12))
        
        tk.Label(step2_header_inner, text="Seleccionar Columnas de Teléfono",
                font=('Inter', 14, 'bold'),
                bg='#f5f5f5', fg=self.colors['text']).pack(side=tk.LEFT)
        
        step2_arrow = tk.Label(step2_header_inner, text="▼", font=('Inter', 16, 'bold'),
                              bg='#f5f5f5', fg=self.colors['green'])
        step2_arrow.pack(side=tk.RIGHT, padx=10)
        
        step2_content = tk.Frame(step2, bg='white')
        
        phone_box = tk.Frame(step2_content, bg='#fff9e6', relief=tk.SOLID, bd=1)
        phone_box.pack(fill=tk.X, pady=(10, 10))
        
        tk.Label(phone_box, text="Selecciona qué columnas de teléfono usar para el envío:",
                font=('Inter', 11),
                bg='#fff9e6', fg='#856404').pack(anchor='w', padx=20, pady=(15, 10))
        
        self.phone_vars = {}
        for idx, phone_col in enumerate(self.phone_columns):
            var = tk.BooleanVar(value=(idx == 0))
            self.phone_vars[phone_col] = var
            
            cb_frame = tk.Frame(phone_box, bg='#fff9e6')
            cb_frame.pack(anchor='w', padx=20, pady=4)
            
            cb = tk.Checkbutton(cb_frame, text=phone_col, variable=var,
                               font=('Inter', 11, 'bold'),
                               bg='#fff9e6', fg='#856404',
                               selectcolor='#fff9e6',
                               activebackground='#fff9e6')
            cb.pack(side=tk.LEFT)
        
        tk.Label(phone_box, text=" ", bg='#fff9e6').pack(pady=5)
        
        def toggle_step2(event=None):
            if step2_content.winfo_ismapped():
                step2_content.pack_forget()
                step2_arrow.config(text="▼")
            else:
                step2_content.pack(fill=tk.X, pady=(0, 0), after=step2_toggle)
                step2_arrow.config(text="▲")
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            return "break"
        
        step2_toggle.bind('<Button-1>', toggle_step2)
        for widget in step2_header_inner.winfo_children():
            widget.bind('<Button-1>', toggle_step2)
        
        tk.Frame(step2, bg='#e0e0e0', height=2).pack(fill=tk.X, pady=15)
        
        # PASO 3: Columnas (desplegable mejorado)
        step3 = tk.Frame(scrollable_frame, bg='white')
        step3.pack(fill=tk.X, padx=30, pady=10)
        
        step3_toggle = tk.Frame(step3, bg='#f5f5f5', relief=tk.FLAT, bd=0, cursor='hand2')
        step3_toggle.pack(fill=tk.X)
        
        step3_header_inner = tk.Frame(step3_toggle, bg='#f5f5f5')
        step3_header_inner.pack(fill=tk.X, padx=15, pady=12)
        
        tk.Label(step3_header_inner, text="2", font=('Inter', 18, 'bold'),
                bg='#e8eaed', fg='#202124', width=2, height=1).pack(side=tk.LEFT, padx=(0, 12))
        
        tk.Label(step3_header_inner, text="Seleccionar Columnas para el Mensaje",
                font=('Inter', 14, 'bold'),
                bg='#f5f5f5', fg=self.colors['text']).pack(side=tk.LEFT)
        
        step3_arrow = tk.Label(step3_header_inner, text="▼", font=('Inter', 16, 'bold'),
                              bg='#f5f5f5', fg=self.colors['orange'])
        step3_arrow.pack(side=tk.RIGHT, padx=10)
        
        step3_content = tk.Frame(step3, bg='white')
        
        columns_box = tk.Frame(step3_content, bg='#e3f2fd', relief=tk.SOLID, bd=1)
        columns_box.pack(fill=tk.X, pady=(10, 10))
        
        tk.Label(columns_box, text="Selecciona las columnas que quieres usar en el mensaje:",
                font=('Inter', 11),
                bg='#e3f2fd', fg='#0d47a1').pack(anchor='w', padx=20, pady=(15, 10))
        
        self.column_vars = {}
        
        cols_grid = tk.Frame(columns_box, bg='#e3f2fd')
        cols_grid.pack(fill=tk.X, padx=20, pady=(0, 15))
        
        col_count = 0
        row_count = 0
        for col in self.columns:
            if col and col not in self.phone_columns:
                var = tk.BooleanVar(value=False)
                self.column_vars[col] = var
                
                cb = tk.Checkbutton(cols_grid, text=col, variable=var,
                                   font=('Inter', 10),
                                   bg='#e3f2fd', fg='#0d47a1',
                                   selectcolor='#e3f2fd',
                                   activebackground='#e3f2fd')
                cb.grid(row=row_count, column=col_count, sticky='w', padx=10, pady=4)
                
                col_count += 1
                if col_count >= 3:
                    col_count = 0
                    row_count += 1
        
        def toggle_step3(event=None):
            if step3_content.winfo_ismapped():
                step3_content.pack_forget()
                step3_arrow.config(text="▼")
            else:
                step3_content.pack(fill=tk.X, pady=(0, 0), after=step3_toggle)
                step3_arrow.config(text="▲")
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            return "break"
        
        step3_toggle.bind('<Button-1>', toggle_step3)
        for widget in step3_header_inner.winfo_children():
            widget.bind('<Button-1>', toggle_step3)
        
        tk.Frame(step3, bg='#e0e0e0', height=2).pack(fill=tk.X, pady=15)
        
        # PASO 4: Mensaje (desplegable mejorado)
        step4 = tk.Frame(scrollable_frame, bg='white')
        step4.pack(fill=tk.X, padx=30, pady=10)
        
        step4_toggle = tk.Frame(step4, bg='#f5f5f5', relief=tk.FLAT, bd=0, cursor='hand2')
        step4_toggle.pack(fill=tk.X)
        
        step4_header_inner = tk.Frame(step4_toggle, bg='#f5f5f5')
        step4_header_inner.pack(fill=tk.X, padx=15, pady=12)
        
        tk.Label(step4_header_inner, text="3", font=('Inter', 18, 'bold'),
                bg='#e8eaed', fg='#202124', width=2, height=1).pack(side=tk.LEFT, padx=(0, 12))
        
        tk.Label(step4_header_inner, text="Plantilla de Mensaje",
                font=('Inter', 14, 'bold'),
                bg='#f5f5f5', fg=self.colors['text']).pack(side=tk.LEFT)
        
        step4_arrow = tk.Label(step4_header_inner, text="▼", font=('Inter', 16, 'bold'),
                              bg='#f5f5f5', fg=self.colors['blue'])
        step4_arrow.pack(side=tk.RIGHT, padx=10)
        
        step4_content = tk.Frame(step4, bg='white')
        
        message_box = tk.Frame(step4_content, bg='#f0f8ff', relief=tk.SOLID, bd=1)
        message_box.pack(fill=tk.X, pady=(10, 10))
        
        tk.Label(message_box, text="Escribe tu mensaje usando {NombreColumna} para insertar valores:",
                font=('Inter', 11),
                bg='#f0f8ff', fg='#0d47a1').pack(anchor='w', padx=20, pady=(15, 10))
        
        buttons_frame = tk.Frame(message_box, bg='#f0f8ff')
        buttons_frame.pack(fill=tk.X, padx=20, pady=(0, 10))
        
        tk.Label(buttons_frame, text="Haz clic para insertar:",
                font=('Inter', 10, 'italic'),
                bg='#f0f8ff', fg='#666').pack(anchor='w', pady=(0, 5))
        
        buttons_container = tk.Frame(buttons_frame, bg='#f0f8ff')
        buttons_container.pack(fill=tk.X)
        
        message_text = scrolledtext.ScrolledText(message_box, height=6, 
                                                 font=('Inter', 11),
                                                 relief=tk.SOLID, bd=1, wrap=tk.WORD)
        message_text.pack(fill=tk.BOTH, padx=20, pady=(10, 10))
        
        # PREVISUALIZACIÓN EN TIEMPO REAL
        preview_frame = tk.Frame(message_box, bg='#fff3cd', relief=tk.SOLID, bd=1)
        preview_frame.pack(fill=tk.BOTH, padx=20, pady=(10, 15))
        
        tk.Label(preview_frame, text="👁️ Previsualización del mensaje:",
                font=('Inter', 10, 'bold'),
                bg='#fff3cd', fg='#856404').pack(anchor='w', padx=10, pady=(8, 5))
        
        preview_text = tk.Text(preview_frame, height=4, 
                              font=('Inter', 10),
                              bg='#fffbf0', fg='#333',
                              relief=tk.FLAT, wrap=tk.WORD,
                              state=tk.DISABLED)
        preview_text.pack(fill=tk.BOTH, padx=10, pady=(0, 8))
        
        def update_preview(*args):
            try:
                # Obtener el mensaje actual
                current_message = message_text.get('1.0', tk.END).strip()
                
                if not current_message:
                    preview_text.config(state=tk.NORMAL)
                    preview_text.delete('1.0', tk.END)
                    preview_text.insert('1.0', '(Escribe tu mensaje arriba para ver la previsualización)')
                    preview_text.config(state=tk.DISABLED)
                    return
                
                # Usar la primera fila de datos como ejemplo
                if self.raw_data:
                    example_row = self.raw_data[0]
                    preview_message = current_message
                    
                    # Reemplazar cada placeholder con datos reales
                    for col in self.columns:
                        placeholder = f"{{{col}}}"
                        if placeholder in preview_message:
                            value = example_row.get(col, '')
                            if value is None:
                                value = ''
                            
                            # Formatear como peso si es necesario
                            if '$ Hist.' in col or '$ Asig.' in col:
                                try:
                                    num_value = float(str(value).replace(',', '').replace('$', '').strip())
                                    value = f"${num_value:,.2f}"
                                except:
                                    value = str(value)
                            else:
                                value = str(value)
                            
                            preview_message = preview_message.replace(placeholder, value)
                    
                    preview_text.config(state=tk.NORMAL)
                    preview_text.delete('1.0', tk.END)
                    preview_text.insert('1.0', preview_message)
                    preview_text.config(state=tk.DISABLED)
                else:
                    preview_text.config(state=tk.NORMAL)
                    preview_text.delete('1.0', tk.END)
                    preview_text.insert('1.0', '(No hay datos para previsualizar)')
                    preview_text.config(state=tk.DISABLED)
            except Exception as e:
                pass
        
        # Actualizar previsualización cuando cambia el texto
        message_text.bind('<KeyRelease>', update_preview)
        message_text.bind('<ButtonRelease>', update_preview)
        
        # Previsualización inicial
        update_preview()
        
        def update_buttons():
            for widget in buttons_container.winfo_children():
                widget.destroy()
            
            selected = [col for col, var in self.column_vars.items() if var.get()]
            
            if not selected:
                tk.Label(buttons_container, text="(Selecciona columnas en el Paso 3)",
                        font=('Inter', 10, 'italic'),
                        bg='#f0f8ff', fg='#999').pack(anchor='w')
                return
            
            def insert_field(field_name):
                message_text.insert(tk.INSERT, f"{{{field_name}}}")
                message_text.focus()
            
            btn_col = 0
            btn_row = 0
            for col in selected:
                btn = tk.Button(buttons_container, text=col,
                               command=lambda c=col: insert_field(c),
                               bg=self.colors['blue'], fg='white',
                               font=('Inter', 9, 'bold'),
                               relief=tk.FLAT, cursor='hand2',
                               padx=12, pady=6)
                btn.grid(row=btn_row, column=btn_col, padx=3, pady=3, sticky='ew')
                
                btn_col += 1
                if btn_col >= 4:
                    btn_col = 0
                    btn_row += 1
        
        for var in self.column_vars.values():
            var.trace('w', lambda *args: update_buttons())
        
        update_buttons()
        
        def toggle_step4(event=None):
            if step4_content.winfo_ismapped():
                step4_content.pack_forget()
                step4_arrow.config(text="▼")
            else:
                step4_content.pack(fill=tk.X, pady=(0, 0), after=step4_toggle)
                step4_arrow.config(text="▲")
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
            return "break"
        
        step4_toggle.bind('<Button-1>', toggle_step4)
        for widget in step4_header_inner.winfo_children():
            widget.bind('<Button-1>', toggle_step4)
        
        # Botones finales
        button_frame = tk.Frame(scrollable_frame, bg='white')
        button_frame.pack(fill=tk.X, padx=30, pady=(20, 30))
        
        def process_and_close():
            selected_phones = [col for col, var in self.phone_vars.items() if var.get()]
            
            if not selected_phones:
                messagebox.showwarning("Advertencia", "Selecciona al menos una columna de teléfono")
                return
            
            selected = [col for col, var in self.column_vars.items() if var.get()]
            
            if not selected:
                messagebox.showwarning("Advertencia", "Selecciona al menos una columna para el mensaje")
                return
            
            message_template = message_text.get("1.0", tk.END).strip()
            
            if not message_template:
                messagebox.showwarning("Advertencia", "Escribe una plantilla de mensaje")
                return
            
            self.log("⚙️ Procesando datos...", 'info')
            self.process_excel_data(selected, message_template, selected_phones)
            
            proc_window.destroy()
        
        btn_process = tk.Button(button_frame, text="✓ Procesar y Generar URLs",
                 command=process_and_close,
                 bg=self.colors['green'], fg='white',
                 font=('Inter', 13, 'bold'),
                 relief=tk.FLAT, cursor='hand2',
                 activebackground='#17A34A',
                 padx=30, pady=15)
        btn_process.pack(side=tk.LEFT, padx=5)
        
        btn_cancel = tk.Button(button_frame, text="✗ Cancelar",
                 command=proc_window.destroy,
                 bg='#EA4335', fg='white',
                 font=('Inter', 13, 'bold'),
                 relief=tk.FLAT, cursor='hand2',
                 activebackground='#D33426',
                 padx=30, pady=15)
        btn_cancel.pack(side=tk.LEFT, padx=5)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def process_excel_data(self, selected_columns, message_template, selected_phones):
        """Procesar datos y generar URLs"""
        processed_rows = []
        
        for row in self.raw_data:
            phone_numbers = []
            
            for phone_col in selected_phones:
                phone_value = str(row.get(phone_col, '')) if row.get(phone_col) else ''
                numbers = [num.strip() for num in phone_value.split('-') if num.strip()]
                phone_numbers.extend(numbers)
            
            if not phone_numbers:
                continue
            
            for phone in phone_numbers:
                if phone and phone.strip():
                    message = message_template
                    for col in selected_columns:
                        placeholder = f"{{{col}}}"
                        value = row.get(col, '')
                        if value is None:
                            value = ''
                        
                        # Formatear como peso si la columna contiene "$ Hist." o "$ Asig."
                        if '$ Hist.' in col or '$ Asig.' in col:
                            try:
                                # Convertir a número y formatear como peso
                                num_value = float(str(value).replace(',', '').replace('$', '').strip())
                                value = f"${num_value:,.2f}"
                            except:
                                value = str(value)
                        else:
                            value = str(value)
                        
                        message = message.replace(placeholder, value)
                    
                    phone_clean = phone.strip()
                    # Codificar mensaje preservando emojis (safe='' para codificar todo excepto caracteres seguros)
                    encoded_message = urllib.parse.quote(message, safe='')
                    whatsapp_url = f"https://wa.me/549{phone_clean}?text={encoded_message}"
                    
                    processed_rows.append(whatsapp_url)
        
        self.links = processed_rows
        self.total_messages = len(self.links)
        self.update_stats()

        self.log(f"✓ {len(self.links)} URLs de WhatsApp generados", 'success')

        if not self.manual_mode:
            self.save_processed_excel()
    
    def save_processed_excel(self):
        """Guardar Excel con URLs"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "URLs WhatsApp"
            
            ws['A1'] = 'URL'
            
            for idx, url in enumerate(self.links, start=2):
                ws[f'A{idx}'] = url
            
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar Excel Procesado"
            )
            
            if output_path:
                wb.save(output_path)
                self.log(f"✓ Excel guardado: {os.path.basename(output_path)}", 'success')
                messagebox.showinfo("Éxito", f"Excel procesado guardado correctamente\n{len(self.links)} URLs listos para enviar")
        
        except Exception as e:
            self.log(f"✗ Error al guardar Excel: {e}", 'error')
            messagebox.showerror("Error", f"Error al guardar Excel: {e}")
            
    def start_sending(self):
        """Iniciar envío"""
        if not self.adb_path.get() or not os.path.exists(self.adb_path.get()):
            messagebox.showerror("Error", "ADB no encontrado")
            return
        if not self.devices:
            messagebox.showerror("Error", "Paso 1: Detecta dispositivos")
            return
        if not self.links:
            messagebox.showerror("Error", "Pasos 2/3: Carga datos desde Excel o lista manual")
            return
        if self.is_running:
            return
            
        if not messagebox.askyesno("Confirmar",
            f"¿Iniciar envío de {len(self.links)} mensajes?"):
            return
            
        self.is_running = True
        self.is_paused = False
        self.should_stop = False
        self.sent_count = 0
        self.failed_count = 0
        self.current_index = 0
        self.start_time = datetime.now()
        
        self.btn_start.config(state=tk.DISABLED)
        self.btn_pause.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.NORMAL)
        
        threading.Thread(target=self.send_thread, daemon=True).start()
        
    def pause_sending(self):
        """Pausar/Reanudar"""
        with self.pause_lock:
            if self.is_paused:
                self.is_paused = False
                self.btn_pause.config(text="⏸  PAUSAR")
                self.log("▶ Reanudado", 'success')
            else:
                self.is_paused = True
                self.btn_pause.config(text="▶  REANUDAR")
                self.log("⏸ Pausado", 'warning')
                
    def stop_sending(self):
        """Cancelar"""
        if messagebox.askyesno("Confirmar", "¿Cancelar el envío?"):
            self.should_stop = True
            self.log("⏹ Cancelando...", 'warning')
            
    def send_thread(self):
        """Thread de envío"""
        try:
            self.log("═" * 50, 'info')
            self.log("🚀 INICIANDO ENVÍO", 'success')
            self.log("═" * 50, 'info')

            pkg = "com.whatsapp.w4b"
            chrome = "com.android.chrome/com.google.android.apps.chrome.Main"
            idx = 0

            for device in self.devices:
                if self.should_stop:
                    break
                self.close_all_apps(device)

            if self.should_stop:
                self.log("⚠ Envío cancelado", 'warning')
                return

            self.log("🕒 Esperando 3s antes de iniciar el envío...", 'info')
            while self.is_paused and not self.should_stop:
                time.sleep(0.1)
            if self.should_stop:
                self.log("⚠ Envío cancelado", 'warning')
                return
            time.sleep(3)
            if self.should_stop:
                self.log("⚠ Envío cancelado", 'warning')
                return

            for i, link in enumerate(self.links, 1):
                while self.is_paused and not self.should_stop:
                    time.sleep(0.1)
                if self.should_stop:
                    self.log("⚠ Envío cancelado", 'warning')
                    break
                    
                self.current_index = i
                device = self.devices[idx]
                idx = (idx + 1) % len(self.devices)

                self.close_all_apps(device)

                while self.is_paused and not self.should_stop:
                    time.sleep(0.1)
                if self.should_stop:
                    self.log("⚠ Envío cancelado", 'warning')
                    break

                if self.send_msg(device, link, i, len(self.links), pkg, chrome):
                    self.sent_count += 1
                else:
                    self.failed_count += 1
                    
                self.update_stats()
                
                if i < len(self.links) and not self.should_stop:
                    delay = random.uniform(self.delay_min.get(),
                                          self.delay_max.get())
                    self.log(f"⏳ Esperando {delay:.1f}s...", 'info')
                    
                    elapsed = 0
                    while elapsed < delay and not self.should_stop:
                        while self.is_paused and not self.should_stop:
                            time.sleep(0.1)
                        time.sleep(0.1)
                        elapsed += 0.1
                        
            self.log("═" * 50, 'info')
            self.log("✅ ENVÍO FINALIZADO", 'success')
            self.log(f"Enviados: {self.sent_count} | Fallidos: {self.failed_count}", 'info')
            
            messagebox.showinfo("Completado",
                f"Enviados: {self.sent_count}\nFallidos: {self.failed_count}")
        finally:
            self.is_running = False
            self.btn_start.config(state=tk.NORMAL)
            self.btn_pause.config(state=tk.DISABLED)
            self.btn_stop.config(state=tk.DISABLED)

    def send_msg(self, device, link, i, total, pkg, chrome):
        """Enviar mensaje - Abre Google e inyecta URL"""
        try:
            num = link.split('wa.me/')[1].split('?')[0] if 'wa.me/' in link else "?"
            self.log(f"📱 {i}/{total} → {num}", 'info')
            
            adb = self.adb_path.get()
            
            # Cerrar WhatsApp primero
            subprocess.run([adb, '-s', device, 'shell', 'am', 'force-stop', pkg], 
                          capture_output=True, timeout=10)
            time.sleep(1)
            
            # Abrir Google app e inyectar URL
            self.log("🔗 Abriendo Google e inyectando URL...", 'info')
            
            # Usar monkey para abrir Google con el URL
            cmd = f'monkey -p com.google.android.googlequicksearchbox -c android.intent.category.LAUNCHER 1 && sleep 1 && am start -a android.intent.action.VIEW -d "{link}"'
            subprocess.run([adb, '-s', device, 'shell', cmd], 
                          capture_output=True, timeout=15, shell=False)
            
            time.sleep(self.wait_after_open.get())
            
            # Primer Enter (abrir chat en WhatsApp)
            subprocess.run([adb, '-s', device, 'shell', 'input', 'keyevent', '66'], 
                          capture_output=True, timeout=10)
            time.sleep(self.wait_after_first_enter.get())
            
            # Segundo Enter (enviar mensaje)
            subprocess.run([adb, '-s', device, 'shell', 'input', 'keyevent', '66'], 
                          capture_output=True, timeout=10)
            time.sleep(1)
            
            self.log("✅ ENVIADO", 'success')
            return True
        except subprocess.TimeoutExpired:
            self.log("❌ ERROR: Timeout", 'error')
            return False
        except Exception as e:
            self.log(f"❌ ERROR: {e}", 'error')
            return False

    def close_all_apps(self, device):
        """Cerrar aplicaciones antes de iniciar el envío"""
        adb = self.adb_path.get()
        if not adb:
            self.log("⚠ No se puede cerrar apps: ADB no configurado", 'warning')
            return

        self.log(f"🧹 Cerrando WhatsApp y Google en {device}...", 'info')

        targets = [
            ("WhatsApp Business", "com.whatsapp.w4b"),
            ("WhatsApp", "com.whatsapp"),
            ("Google", "com.google.android.googlequicksearchbox"),
        ]

        had_error = False

        for label, package in targets:
            try:
                result = subprocess.run(
                    [adb, '-s', device, 'shell', 'am', 'force-stop', package],
                    capture_output=True,
                    text=True,
                    timeout=10
                )
                if result.returncode != 0:
                    had_error = True
                    error_msg = result.stderr.strip() or result.stdout.strip() or "Error desconocido"
                    self.log(
                        f"⚠ No se pudo cerrar {label} ({package}) en {device}: {error_msg}",
                        'warning'
                    )
            except subprocess.TimeoutExpired:
                had_error = True
                self.log(f"❌ Timeout al forzar cierre de {label} ({package}) en {device}", 'error')
            except Exception as exc:
                had_error = True
                self.log(f"❌ Error al forzar cierre de {label} ({package}) en {device}: {exc}", 'error')

        if not had_error:
            self.log(f"✅ Apps cerradas correctamente en {device}", 'success')


def main():
    root = tk.Tk()
    Hermes(root)
    root.mainloop()

if __name__ == "__main__":
    main()

